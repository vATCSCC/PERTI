"""Inspect actual Excel formulas from the VATUSA Event Statistics workbook."""

import openpyxl
from pathlib import Path

# The xlsx file (Event List, TMR data)
EVENTS_XLSX = Path(r"C:\Users\jerem.DESKTOP-T926IG8\OneDrive - Virtual Air Traffic Control System Command Center\Documents - Virtual Air Traffic Control System Command Center\DCC\VATUSA Events Data.xlsx")

# The xlsb file (hourly traffic data) - note: pyxlsb cannot read formulas, only values
STATS_XLSB = Path(r"C:\Users\jerem.DESKTOP-T926IG8\OneDrive - Virtual Air Traffic Control System Command Center\Documents - Virtual Air Traffic Control System Command Center\DCC\VATUSA Event Statistics.xlsb")

def inspect_formulas():
    print(f"Reading: {EVENTS_XLSX}")

    # Load with data_only=False to see formulas
    wb = openpyxl.load_workbook(EVENTS_XLSX, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}")

        # Get headers from row 1
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                headers[col] = str(cell.value)

        print(f"Columns: {list(headers.values())[:15]}...")  # First 15 columns

        # Check first few data rows for formulas
        formula_cols = {}
        for row in range(2, min(6, ws.max_row + 1)):  # Check rows 2-5
            for col, header in headers.items():
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    if col not in formula_cols:
                        formula_cols[col] = {
                            'header': header,
                            'formula': cell.value,
                            'row': row
                        }

        if formula_cols:
            print(f"\nFormula columns found ({len(formula_cols)} total):")
            for col, info in sorted(formula_cols.items()):
                # Simplify formula for display
                formula = info['formula']
                # Remove table reference prefix for readability
                formula = formula.replace("Table13456781011121346536061718386879899[[#This Row],", "[")
                formula = formula.replace("Table13456781011121346536061718386879899[", "[")
                print(f"  {col:3}. {info['header'][:35]:35} = {formula[:80]}")
        else:
            print("\nNo formulas found (values only)")

if __name__ == "__main__":
    inspect_formulas()

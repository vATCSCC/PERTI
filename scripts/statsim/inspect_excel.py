#!/usr/bin/env python3
"""
Quick script to inspect Excel files and show their structure.
Run this to see column names and sample data.

Usage:
    python inspect_excel.py

Requirements:
    pip install openpyxl pyxlsb
"""

import sys
from pathlib import Path

# Try to import Excel libraries
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

try:
    from pyxlsb import open_workbook as open_xlsb
except ImportError:
    print("Installing pyxlsb...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pyxlsb"])
    from pyxlsb import open_workbook as open_xlsb


def inspect_xlsx(filepath: str):
    """Inspect .xlsx file structure."""
    print(f"\n{'='*60}")
    print(f"FILE: {filepath}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(max_row=6, values_only=True))
        if not rows:
            print("  (empty sheet)")
            continue

        # Headers
        headers = rows[0] if rows else []
        print(f"\nColumns ({len([h for h in headers if h])} non-empty):")
        for i, h in enumerate(headers):
            if h:
                print(f"  [{i}] {h}")

        # Sample data
        print(f"\nSample rows (first 5):")
        for row_num, row in enumerate(rows[1:6], start=2):
            print(f"  Row {row_num}: {row[:10]}{'...' if len(row) > 10 else ''}")

    wb.close()


def inspect_xlsb(filepath: str):
    """Inspect .xlsb file structure."""
    print(f"\n{'='*60}")
    print(f"FILE: {filepath}")
    print(f"{'='*60}")

    with open_xlsb(filepath) as wb:
        for sheet_name in wb.sheets:
            print(f"\n--- Sheet: {sheet_name} ---")

            with wb.get_sheet(sheet_name) as sheet:
                rows = []
                for row in sheet.rows():
                    rows.append([item.v for item in row])
                    if len(rows) >= 6:
                        break

                if not rows:
                    print("  (empty sheet)")
                    continue

                # Headers
                headers = rows[0] if rows else []
                print(f"\nColumns ({len([h for h in headers if h])} non-empty):")
                for i, h in enumerate(headers):
                    if h:
                        print(f"  [{i}] {h}")

                # Sample data
                print(f"\nSample rows (first 5):")
                for row_num, row in enumerate(rows[1:6], start=2):
                    print(f"  Row {row_num}: {row[:10]}{'...' if len(row) > 10 else ''}")


def main():
    # Your Excel files
    files = [
        r"C:\Users\jerem.DESKTOP-T926IG8\OneDrive - Virtual Air Traffic Control System Command Center\Documents - Virtual Air Traffic Control System Command Center\DCC\VATUSA Events Data.xlsx",
        r"C:\Users\jerem.DESKTOP-T926IG8\OneDrive - Virtual Air Traffic Control System Command Center\Documents - Virtual Air Traffic Control System Command Center\DCC\VATUSA Event Statistics.xlsb",
    ]

    for filepath in files:
        path = Path(filepath)
        if not path.exists():
            print(f"\nFile not found: {filepath}")
            continue

        try:
            if path.suffix.lower() == '.xlsx':
                inspect_xlsx(filepath)
            elif path.suffix.lower() == '.xlsb':
                inspect_xlsb(filepath)
            else:
                print(f"\nUnknown format: {path.suffix}")
        except Exception as e:
            print(f"\nError reading {filepath}: {e}")

    print("\n" + "="*60)
    print("Copy the output above and share it to show me your data structure.")
    print("="*60)


if __name__ == '__main__':
    main()
